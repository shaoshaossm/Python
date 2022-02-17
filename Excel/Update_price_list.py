import openpyxl

wb = openpyxl.load_workbook('电子产品价格表.xlsx')
sheet = wb['Sheet1']
PRICE_UPDATES = {'苹果': 3,
                 '香蕉': 4,
                 '橘子': 5
                 }
for rowNum in range(2, sheet.max_row + 1):
    productName = sheet.cell(row=rowNum, column=1).value
    if productName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[productName]
wb.save('updateProductsSales.xlsx')
