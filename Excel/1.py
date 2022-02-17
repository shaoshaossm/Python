import openpyxl

print(openpyxl.__version__)
# 用openpyxl读取excel表格
wb = openpyxl.load_workbook('信息表.xlsx')
print(wb)
# 获取工作蒲sheet表名称
sheet1 = wb.sheetnames
print("sheet表名称:\n", sheet1)
# 获取指定sheet对象
sheet = wb['基本信息']
print(sheet)
# 获取活动表(目前活动的表)
print(wb.active)
# 从表中获取单元格
cell = sheet['A4']  # 创建一个cell对象
# 参数
print(cell.value)  # value:cell中存储的值
print(cell.row)  # 行索引
print(cell.column)  # 列索引
print(cell.coordinate)  # 坐标
# 用字母指定行列
print(sheet.cell(row=4, column=2).value)
#  遍历一个矩形区域中的所有cell对象

for cell_row in sheet['A3':'D8']:
    for cell in cell_row:
        print(cell.coordinate, cell.value)
# 访问特定行或列的单元格的值,利用worksheet对象的rows和column属性
print(list(sheet.columns))
for cell in list(sheet.columns)[0]:
    print(cell.value)
# 获取工作表中最大行和最大列的数量
print('--------------------------')
print(sheet.max_row, sheet.max_column)
