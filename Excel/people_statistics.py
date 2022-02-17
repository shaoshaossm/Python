import openpyxl, pprint

print('opening workbook...')
wb = openpyxl.load_workbook('人口统计简易表.xlsx')
sheet = wb['人口统计表']
print(wb.sheetnames)
# countryData将包含计算的每个县的总人口和普查区数目
countyData = {}
print('Reading rows...')

for row in range(2, sheet.max_row + 1):
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value
    countyData.setdefault(state, {})
    countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})
    countyData[state][county]['tracts']+=1
    countyData[state][county]['pop']+=int(pop)
print('Writing results...')
resultFile = open('census2022.py','w')
# 将countyData数据写入到addData字典中
resultFile.write('allData = '+pprint.pformat(countyData))
resultFile.close()
print('Done')


