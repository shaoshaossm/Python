from openpyxl.styles import Font, PatternFill, Side, Border, Alignment
import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = '字体'
# italic: 斜体 underline: 下划线 b: 加粗 color: RGB值
sheet['A1'].font = Font(name='楷体', color='8470FF', size=12, italic=True, underline='single', b=True)
# cell填充色
sheet['A2'].fill = PatternFill(patternType='solid', fgColor='8470FF')
# 设置边框样式
sheet['A4'] = '效果1'
sheet['A5'] = '效果2'
s1 = Side(style='thin', color='8470FF')
s2 = Side(style='double', color='8470FF')
sheet['A4'].border = Border(top=s1)
sheet['A5'].border = Border(top=s2, bottom=s1, left=s2, right=s1)

sheet['B1'] = '效果1'
sheet['B2'] = '效果2'
sheet['B3'] = '效果3'
B1 = sheet['B1'].alignment = Alignment(horizontal='right')
B2 = sheet['B2'].alignment = Alignment(horizontal='center')
B3 = sheet['B3'].alignment = Alignment(horizontal='general')






wb.save('styles.xlsx')
